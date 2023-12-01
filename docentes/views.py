from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl import Workbook

import docentes
from docentes.forms import DocenteFormulario
from docentes.models import Docente

# Create your views here.
DocenteFormulario = modelform_factory(Docente, exclude=['activo',])
# Create your views here.
def agregar_docente(request):
    pagina = loader.get_template('docentes/agregar.html')
    if request.method == 'GET':
        formulario = DocenteFormulario
    elif request.method == 'POST':
        formulario = DocenteFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')

    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def modificar_docente(request, id):
    pagina = loader.get_template('docentes/modificar.html')
    docente = get_object_or_404(Docente, pk=id)
    if request.method == 'GET':
        formulario = DocenteFormulario(instance=docente)
    elif request.method == 'POST':
        formulario = DocenteFormulario(request.POST, instance=docente )
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_docente(request, id):
    #docente = Docente.objects.get(pk=id)
    docente = get_object_or_404(Docente, pk=id)
    datos ={'docente': docente}
    print(docente)
    pagina = loader.get_template('docentes/ver.html')
    return HttpResponse(pagina.render(datos, request))

def eliminar_docente(request, id):
    docente = get_object_or_404(Docente, pk=id)
    if docente:
        docente.delete()
        return redirect('inicio')

# Create your views here.
# Nuestra clase hereda de la vista genérica TemplateView

def generar_reporte(request, *args, **kwargs):
    # Obtenemos todas las personas de nuestra base de datos
    docentes = Docente.objects.order_by('apellido','nombre')

    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE PERSONAS'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'Id'
    ws['C3'] = 'NOMBRE'
    ws['D3'] = 'APELLIDO'
    ws['E3'] = 'CEDULA'
    ws['F3'] = 'EMAIL'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for docente in docentes:
        ws.cell(row=cont, column=2).value = docente.id
        ws.cell(row=cont, column=3).value = docente.nombre
        ws.cell(row=cont, column=4).value = docente.apellido
        ws.cell(row=cont, column=5).value = docente.cedula
        ws.cell(row=cont, column=5).value = docente.email
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReportePersonasExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response